import logging

import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\1030968\\PycharmProjects\\PythonSelfFramework\\testData\\PytestData.xlsx")
sheet1 = workbook.active
rows = sheet1.max_row
print(rows)
columns = sheet1.max_column
dictionaryData = {}
list_of_data = []

class HomePageData:
    test_HomePage_data = [{"FirstName":"Deepti", "Email":"deeps15th@gmail.com", "Password":"Hi", "Gender":"Female",
                           "AdditionalText":"Hello"},
                        {"FirstName":"Ritesh", "Email":"ritesh123@gmail.com", "Password":"Message", "Gender":"Male",
                         "AdditionalText":"Good Morning"}]

    @staticmethod
    def get_values_from_excel():
        for i in range(1, rows + 1):
            # if sheet1.cell(row=i, column=1).value == "TestCase2":
            for j in range(2, columns + 1):
                dictionaryData[sheet1.cell(row=1, column=j).value] = sheet1.cell(row=i, column=j).value
            print(dictionaryData)
            list_of_data.append(dictionaryData)
        return list_of_data

    # def getLogger(self):
    #     logger = logging.getLogger()
    #     filehandler = logging.FileHandler("Hello.log")
    #     formatter = logging.Formatter("")
    #     filehandler.setFormatter(formatter)
    #     logger.addHandler(filehandler)
    #     logger.setLevel(logging.DEBUG)