import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\1030968\\PycharmProjects\\PythonSelfFramework\\testData\\PytestData.xlsx")
sheet1 = workbook.active
cell = sheet1.cell(row=1, column=4)
print(cell.value)
sheet1.cell(row=2, column=2).value = "Shyam"
workbook.save("PytestData.xls")
print(sheet1.cell(row=2, column=2).value)
rows = sheet1.max_row
print(rows)
columns = sheet1.max_column
print(columns)
print(sheet1['B1'].value)
dictionaryData = {}
list_of_data = []


# class ExcelData:
#     @staticmethod
def get_values_from_excel():
    for i in range(1, rows+1):
        #if sheet1.cell(row=i, column=1).value == "TestCase2":
        for j in range(2, columns+1):
            dictionaryData[sheet1.cell(row=1, column=j).value]=sheet1.cell(row=i, column=j).value
        print(dictionaryData)
        list_of_data.append(dictionaryData)
    return list_of_data


print(get_values_from_excel())

